import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from datetime import datetime

# ========================= CONFIG =========================
st.set_page_config(page_title="Générateur PRO + Fusion Multi-Listes", layout="centered", page_icon="🎉")

# ========================= CSS =========================
st.markdown("""
<style>
    .big-title {font-size: 46px !important; text-align: center; color: #FF6B6B; font-weight: bold;}
    .subtitle {font-size: 22px; text-align: center; color: #555; margin-bottom: 30px;}
    .feature-box {background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color:white; padding:15px; border-radius:12px; margin:10px 0; text-align:center; font-weight:bold;}
    .stButton>button {background: linear-gradient(45deg, #FF6B6B, #FF8E53); color:white; border:none; border-radius:15px; height:60px; font-size:20px; font-weight:bold;}
    .stDownloadButton>button {background:#28a745; color:white; font-size:18px; padding:15px;}
</style>
""", unsafe_allow_html=True)

st.markdown("<h1 class='big-title'>🎉 Générateur PRO + Fusion Multi-Listes</h1>", unsafe_allow_html=True)
st.markdown("<p class='subtitle'>Le plus puissant outil de gestion d'invités 2025</p>", unsafe_allow_html=True)

# ========================= MODE CHOICE =========================
mode = st.radio(
    "Mode de fonctionnement :",
    ("Une seule liste (classique)", "🔥 Fusionner plusieurs listes en UN fichier"),
    index=1
)

st.divider()

# ========================= SIDEBAR =========================
with st.sidebar:
    st.header("⚙️ Options avancées")
    couleur_nom = st.color_picker("Couleur NOM", "#E91E63")
    couleur_prenom = st.color_picker("Couleur Prénom(s)", "#9C27B0")
    highlight_vip = st.checkbox("Surligner les VIP en doré", True)
    ajouter_table = st.checkbox("Ajouter colonne 'Table'", True)
    table_start = st.number_input("N° de table de départ", 1, 500, 1, disabled=not ajouter_table)
    gel_en_tete = st.checkbox("Geler la 1ère ligne", True)
    recalculer_rangs = st.checkbox("Recalculer les rangs (fusion)", True)

# ========================= FONCTIONS (DÉFINIES AVANT UTILISATION) =========================
def nettoyer_texte(txt):
    return txt.replace("’", "'").replace("`", "'").replace("´", "'")

def est_remplacement(ligne):
    mots = ["barré", "remplac", "ancienne", "remplace", "supprim", "annul", "absent", "retiré", "rayé"]
    return any(mot in ligne.lower() for mot in mots)

def est_vip(ligne):
    return "★" in ligne or any(k in ligne.lower() for k in ["vip", "président", "ministre", "maire", "dg", "pdg", "directeur", "invité d'honneur"])

def traiter_liste(texte_brut, nom_liste="Liste", debut_rang=1, debut_table=None):
    lignes = [l.strip() for l in texte_brut.split("\n") if l.strip()]
    data = []
    rang = debut_rang
    table = debut_table if (ajouter_table and debut_table is not None) else None

    for ligne_brute in lignes:
        ligne = nettoyer_texte(ligne_brute)
        if not ligne or ligne.startswith("#"):
            continue

        match = re.match(r"^[\s]*(\d+)[\s\.\-\)\]\>\:\-\—\—]+", ligne)
        if match:
            rang = int(match.group(1))

        if est_remplacement(ligne):
            data.append({
                "Rang": rang,
                "Nom": "", "Prénom(s)": "ANCIEN(NE)",
                "Table": table if ajouter_table else "",
                "Remarques": "Remplacé/Barré", "VIP": "Non", "Origine": nom_liste
            })
            if ajouter_table and table is not None:
                table += 1
            rang += 1  # Incrément rang même pour remplacements
            continue

        ligne_nette = re.sub(r"^[\s\d\.\-\)\]\>\:\-\—\—]+", "", ligne).strip()
        vip = "Oui" if est_vip(ligne_nette) else "Non"
        ligne_nette = re.sub(r"★|\(VIP\)", "", ligne_nette, flags=re.IGNORECASE).strip()

        mots = ligne_nette.replace("'", " ").split()
        nom_parts = [m for m in mots if m.isupper() or (len(m)>2 and m[:-1].isupper())]
        prenom_parts = [m for m in mots if m not in nom_parts]

        nom = " ".join(nom_parts).title() if nom_parts else ""
        prenom = " ".join(prenom_parts).title() if prenom_parts else ligne_nette.title()

        if not nom and prenom:
            parts = prenom.split()
            if len(parts) >= 2:
                nom = parts[-1].upper()
                prenom = " ".join(parts[:-1])

        data.append({
            "Rang": rang,
            "Nom": nom,
            "Prénom(s)": prenom,
            "Table": table if ajouter_table else "",
            "Remarques": "VIP ★" if vip == "Oui" else "",
            "VIP": vip,
            "Origine": nom_liste
        })

        rang += 1
        if ajouter_table and table is not None:
            table += 1

    df = pd.DataFrame(data)
    
    # Nettoyage : supprimer colonne Table si désactivée
    if not ajouter_table and "Table" in df.columns:
        df = df.drop(columns=["Table"])
        
    return df

def create_styled_excel(liste_onglets):
    output = BytesIO()
    wb = Workbook()
    
    for idx, (nom_onglet, df) in enumerate(liste_onglets):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = str(nom_onglet)[:31]

        # En-tête
        for c, col in enumerate(df.columns, 1):
            cell = ws.cell(1, c, col)
            cell.font = Font(bold=True, color="FFFFFF", size=13)
            cell.fill = PatternFill("solid", fgColor="1E90FF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            # VIP sécurisé
            vip = False
            if "VIP" in df.columns:
                vip_idx = df.columns.get_loc("VIP")
                if vip_idx < len(row) and row[vip_idx] == "Oui":
                    vip = True

            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(r_idx, c_idx, value)
                if df.columns[c_idx-1] == "Nom":
                    cell.font = Font(color=couleur_nom.replace("#",""), bold=True, size=11)
                if df.columns[c_idx-1] == "Prénom(s)":
                    cell.font = Font(color=couleur_prenom.replace("#",""), italic=True)
                if vip and highlight_vip:
                    cell.fill = PatternFill("solid", fgColor="FFFACD")
                    cell.font = Font(bold=True, color="B8860B")
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        # Ajustement colonnes
        for col in ws.columns:
            max_len = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_len + 3, 50)

        if gel_en_tete:
            ws.freeze_panes = "A2"

    wb.save(output)
    return output.getvalue()

# ========================= MODE CLASSIQUE =========================
if mode == "Une seule liste (classique)":
    col1, col2 = st.columns([3,1])
    with col1:
        texte = st.text_area("📋 Collez votre liste :", height=300)
    with col2:
        fichier = st.file_uploader("📄 Ou un fichier", type=["txt","csv","xlsx","xls"])

    if fichier:
        if fichier.type == "text/plain":
            texte = fichier.read().decode("utf-8", errors="ignore")
        else:
            df_temp = pd.read_excel(fichier) if "excel" in fichier.type else pd.read_csv(fichier)
            texte = "\n".join(df_temp.astype(str).agg(" - ".join, axis=1))

    if st.button("🚀 Générer Excel", type="primary", use_container_width=True):
        if not texte.strip():
            st.error("Aucune donnée détectée")
        else:
            with st.spinner("Traitement en cours..."):
                df = traiter_liste(texte, "Liste Unique", 1, table_start if ajouter_table else None)
                excel = create_styled_excel([("Invités", df)])
                
                st.success("✅ Fichier prêt !")
                st.dataframe(df, use_container_width=True)
                st.download_button(
                    label="📥 Télécharger Excel",
                    data=excel,
                    file_name=f"Liste_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.balloons()

# ========================= MODE FUSION =========================
else:
    st.success("🔥 Chargez plusieurs fichiers → tout sera fusionné automatiquement !")

    uploaded_files = st.file_uploader(
        "Upload multiples (TXT, CSV, Excel)",
        type=["txt", "csv", "xlsx", "xls"],
        accept_multiple_files=True
    )

    texte_multiple = st.text_area(
        "Ou collez plusieurs listes (séparées par --- ou === ou paragraphes vides)",
        height=300
    )

    if st.button("🚀 FUSIONNER TOUT EN UN EXCEL PRO", type="primary", use_container_width=True):
        if not uploaded_files and not texte_multiple.strip():
            st.error("Ajoutez au moins une liste")
        else:
            with st.spinner("Fusion et mise en forme en cours..."):
                toutes_les_df = []
                table_courante = table_start
                rang_courant = 1  # Pour recalcul si besoin

                # 1. Fichiers uploadés
                for idx, file in enumerate(uploaded_files):
                    nom = file.name.split(".")[0][:20] if file.name else f"Fichier {idx+1}"
                    if file.type == "text/plain":
                        content = file.read().decode("utf-8", errors="ignore")
                    else:
                        df_temp = pd.read_excel(file) if "excel" in file.type else pd.read_csv(file)
                        content = "\n".join(df_temp.astype(str).agg(" - ".join, axis=1))

                    df = traiter_liste(content, nom, debut_rang=rang_courant, debut_table=table_courante if ajouter_table else None)
                    if not df.empty:
                        toutes_les_df.append((nom, df))
                        rang_courant += len(df)
                        if ajouter_table:
                            table_courante += len(df) + 5  # Espacement groupes

                # 2. Texte collé
                if texte_multiple.strip():
                    # Split plus robuste : par --- ou === ou paragraphes vides
                    parties = re.split(r"\n(?:[-=]{3,}|#+\s*Liste\s*\d+\s*#*)\n", texte_multiple.strip(), flags=re.IGNORECASE)
                    if len(parties) <= 1:
                        parties = [p.strip() for p in re.split(r"\n{2,}", texte_multiple.strip()) if p.strip()]

                    for i, partie in enumerate(parties):
                        if partie.strip():
                            nom_partie = f"Collée {i+1}"
                            df = traiter_liste(partie, nom_partie, debut_rang=rang_courant, debut_table=table_courante if ajouter_table else None)
                            if not df.empty:
                                toutes_les_df.append((nom_partie, df))
                                rang_courant += len(df)
                                if ajouter_table:
                                    table_courante += len(df) + 3

                if not toutes_les_df:
                    st.error("Aucune liste valide trouvée")
                else:
                    fusion_unique = st.checkbox("Tout fusionner en UN SEUL onglet", value=True)
                    
                    if fusion_unique:
                        df_final = pd.concat([df for _, df in toutes_les_df], ignore_index=True)
                        if recalculer_rangs:
                            df_final["Rang"] = range(1, len(df_final) + 1)
                        onglets = [("TOUS LES INVITÉS", df_final)]
                        nom_fichier = f"TOUS_FUSIONNES_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    else:
                        onglets = toutes_les_df
                        nom_fichier = f"MULTI_ONGLETS_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

                    excel_data = create_styled_excel(onglets)

                    total_invites = len(pd.concat([df for _, df in toutes_les_df], ignore_index=True))
                    st.success(f"✅ {len(toutes_les_df)} liste(s) traitée(s) → {total_invites} invités au total")
                    st.dataframe(pd.concat([df for _, df in toutes_les_df], ignore_index=True), use_container_width=True)

                    st.download_button(
                        label="📥 TÉLÉCHARGER LE FICHIER EXCEL FINAL",
                        data=excel_data,
                        file_name=nom_fichier,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    st.balloons()